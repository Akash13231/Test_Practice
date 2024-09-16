from csv import DictReader
from traceback import format_exc

import openpyxl
Dict = {}
file = openpyxl.load_workbook("C:/Users/abhosage/OneDrive - Capgemini/Desktop/pythondata.xlsx")

sheet = file.active   #to get control on active sheet from Excel file

#first = (sheet.cell(row=2, column=2)).value
# print(sheet.max_row)
# print(sheet.max_column)
# print(first)

for i in range(1, sheet.max_row+1):
    if not sheet.cell(row=i, column=1).value == 'testcase1':
        for j in range(2, sheet.max_column+1):
            Dict[sheet.cell(row=1, column=j).value]=(sheet.cell(row=i, column=j).value)


print(Dict)