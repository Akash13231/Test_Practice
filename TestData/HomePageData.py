
import openpyxl

class HomePageData:
    test_HomePage_data = {'firstname':'akash','email': 'abc@gmail.com','gender': 'Male'}, {'firstname':'Tejashvini', 'email': 'abc@gmail.com', 'gender':'Female'}

    @staticmethod
    def get_exceldata(test_case_name):

        Dict = {}
        file = openpyxl.load_workbook("C:/Users/abhosage/OneDrive - Capgemini/Desktop/pythondata.xlsx")

        sheet = file.active  # to get control on active sheet from Excel file

        for i in range(1, sheet.max_row + 1):
            if not sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    Dict[sheet.cell(row=1, column=j).value] = (sheet.cell(row=i, column=j).value)

        return [Dict]
