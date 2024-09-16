import time

import openpyxl


from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.wait import WebDriverWait



def update_excel(file_path, col_name, froot_name, new_price):
    file = openpyxl.load_workbook(file_path)         #"C:/Users/abhosage/Downloads/download.xlsx"
    sheet = file.active  # to get control on active sheet from Excel file
    Dict = {}

    for i in range(1, sheet.max_column + 1):
        if sheet.cell(row=1, column=i).value == col_name:
            Dict["col"] = i

    for i in range(1, sheet.max_row + 1):
        for j in range(1, sheet.max_column + 1):
            if sheet.cell(row=i, column=j).value == froot_name:
                Dict['row'] = i
    print(Dict)
    sheet.cell(row=Dict["row"], column=Dict["col"]).value = new_price
    file.save(file_path)


froot_name = 'Apple'
file_path = "C:/Users/abhosage/Downloads/download.xlsx"
new_value = '300'
column_name = "price"
driver = webdriver.Chrome()
driver.get('https://rahulshettyacademy.com/upload-download-test/index.html')
driver.maximize_window()
driver.implicitly_wait(6)

driver.find_element(By.ID, 'downloadButton').click()
time.sleep(3)
update_excel(file_path, column_name, froot_name, new_value)

file_input = driver.find_element(By.CSS_SELECTOR, "input[type='file']")
file_input.send_keys(file_path)

wait = WebDriverWait(driver, 10)
toast_locator = (By.CSS_SELECTOR, ".Toastify__toast-body div:nth-child(2)")
wait.until(expected_conditions.visibility_of_element_located(toast_locator))

prise_column = driver.find_element(By.XPATH, "//div[text()='Price']").get_attribute("data-column-id")
actual_prise = driver.find_element(By.XPATH, "//div[text()='"+froot_name+"']/parent::div/parent::div/div[@id='cell-"+prise_column+"-undefined']").text

text = driver.find_element(*toast_locator).text

print(text)

assert actual_prise == new_value

